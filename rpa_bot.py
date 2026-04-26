import os
import re
import time
import json
import queue
import shutil
import csv
import logging
import threading
from datetime import datetime
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from audit_log import audit
from notifier import notify_human_review, load_notification_config

try:
    import requests
    from pypdf import PdfReader
except ImportError:
    print("CRITICAL: Missing dependencies. Run:")
    print("pip install watchdog pypdf requests openpyxl google-generativeai")
    exit(1)

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent
INBOX_DIR      = BASE_DIR / "Inbox"
PROCESSED_DIR  = BASE_DIR / "Processed"
FAILED_DIR     = BASE_DIR / "Failed"
DATABASE_FILE  = BASE_DIR / "Extracted_Database.csv"
EXCEL_FILE     = BASE_DIR / "Extracted_Database.xlsx"
CONFIG_FILE    = BASE_DIR / "config.json"

ENGINE           = 'regex'          # 'regex' | 'gemini' | 'ollama'
GEMINI_API_KEY   = os.environ.get("GEMINI_API_KEY", "")
OLLAMA_URL       = "http://localhost:11434/api/generate"
OLLAMA_MODEL     = "llama3.2"
INTER_FILE_DELAY = 2                # seconds between files (no need to wait for regex)

# ─── SHARED STATE ─────────────────────────────────────────────────────────────
currently_processing = None
processing_lock      = threading.Lock()

# ─── LOGGING ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s %(message)s',
    datefmt='%H:%M:%S'
)
log = logging.getLogger("rpa_bot")

# ─── QUEUE ────────────────────────────────────────────────────────────────────
_file_queue    = queue.Queue()
_worker_thread = None


# ─── CONFIG LOAD / SAVE ───────────────────────────────────────────────────────
def load_config():
    global ENGINE, GEMINI_API_KEY, OLLAMA_MODEL, INTER_FILE_DELAY
    if CONFIG_FILE.exists():
        try:
            cfg = json.loads(CONFIG_FILE.read_text(encoding='utf-8'))
            ENGINE           = cfg.get("engine", ENGINE)
            if not GEMINI_API_KEY:
                GEMINI_API_KEY = cfg.get("api_key", "")
            OLLAMA_MODEL     = cfg.get("ollama_model", OLLAMA_MODEL)
            INTER_FILE_DELAY = cfg.get("inter_file_delay", INTER_FILE_DELAY)
            # Load notification config into notifier module
            load_notification_config(cfg)
            log.info(f"Config loaded  engine={ENGINE}  key={'SET' if GEMINI_API_KEY else 'NOT SET'}")
            audit.log_event("CONFIG_LOADED", engine=ENGINE, inter_file_delay=INTER_FILE_DELAY)
        except Exception as exc:
            log.warning(f"Could not read config.json: {exc}")

def save_config():
    """Write all settings back to config.json, preserving schedule/notifications sections."""
    try:
        # Load existing to preserve schedule + notifications blocks
        existing: dict = {}
        if CONFIG_FILE.exists():
            try:
                existing = json.loads(CONFIG_FILE.read_text(encoding='utf-8'))
            except Exception:
                pass
        existing.update({
            "engine":           ENGINE,
            "api_key":          GEMINI_API_KEY,
            "ollama_model":     OLLAMA_MODEL,
            "inter_file_delay": INTER_FILE_DELAY,
        })
        CONFIG_FILE.write_text(json.dumps(existing, indent=2), encoding='utf-8')
    except Exception as exc:
        log.warning(f"Could not save config.json: {exc}")

def is_configured() -> bool:
    if ENGINE == 'gemini':
        return bool(GEMINI_API_KEY and len(GEMINI_API_KEY) > 10)
    return True  # regex and ollama need no key


# ─── PURE REGEX PARSER (no LLM, works 100% for SuperStore format) ─────────────
def extract_text_from_pdf(file_path: Path) -> str:
    """Extract raw text from a PDF using pypdf."""
    reader = PdfReader(str(file_path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _get(text: str, pattern: str, default: str = "") -> str:
    """Return first capture group of pattern or default."""
    m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return m.group(1).strip() if m else default


def parse_invoice_regex(text: str) -> dict:
    """
    Exact line-by-line parser for the SuperStore invoice format.

    pypdf extracts each PDF text token on its own line. The canonical
    line order is:

        0:  INVOICE
        1:  # <number>
        2:  SuperStore
        3:  Bill To
        4:  :
        5:  <client_name>
        6:  Ship To
        7:  :
        8+: <address lines>   (until date line)
        n:  <Mon DD YYYY>     date
        n+1:<ship mode>
        n+2:$<balance_due>
        n+3:Date
        n+4::
        n+5:Ship Mode
        n+6::
        n+7:Balance Due
        n+8::
        n+9:Item
        n+10:Quantity
        n+11:Rate
        n+12:Amount
    Then for EACH line item:
        desc line
        qty
        $rate
        $amount
        sku_or_category          ← always follows each item
    Then summary block (all on separate lines):
        $subtotal
        $discount  (optional — only if discounted)
        $shipping
        $total
        Subtotal
        :
        [Discount (X%)]
        [:]
        Shipping
        :
        Total
        :
        Notes
        :
        <notes text>
        Terms
        :
        Order ID : <order_id>
    """
    data = {
        "document_type":    "Invoice",
        "reference_number": "",
        "order_id":         "",
        "date":             "",
        "due_date":         "",
        "vendor_name":      "SuperStore",
        "vendor_address":   "",
        "vendor_contact":   "",
        "client_name":      "",
        "client_address":   "",
        "ship_to_address":  "",
        "ship_mode":        "",
        "currency":         "USD",
        "subtotal":         "",
        "discount":         "",
        "shipping_cost":    "",
        "tax":              "",
        "tax_rate":         "",
        "total":            "",
        "balance_due":      "",
        "payment_terms":    "",
        "notes":            "",
        "line_items":       [],
    }

    # Work with stripped, non-empty lines
    lines = [l.strip() for l in text.splitlines()]
    lines = [l for l in lines if l]

    money_re  = re.compile(r'^\$[0-9,]+\.\d{2}$')
    qty_re    = re.compile(r'^\d{1,4}$')
    date_re   = re.compile(r'^[A-Z][a-z]{2}\s+\d{1,2}\s+\d{4}$')
    ship_modes = {"First Class", "Second Class", "Standard Class", "Same Day"}

    n = len(lines)

    # ── 1. Invoice number  (line starting with "# NNNNN")
    for i, line in enumerate(lines):
        m = re.match(r'^#\s*(\d+)$', line)
        if m:
            data["reference_number"] = m.group(1)
            break

    # ── 2. Client name  (line after ":" that follows "Bill To")
    for i, line in enumerate(lines):
        if line == "Bill To" and i + 2 < n and lines[i + 1] == ":":
            data["client_name"] = lines[i + 2]
            break

    # ── 3. Ship To address  (lines after ":" that follows "Ship To", until date line)
    for i, line in enumerate(lines):
        if line == "Ship To" and i + 2 < n and lines[i + 1] == ":":
            addr_parts = []
            j = i + 2
            while j < n and not date_re.match(lines[j]) and lines[j] not in ship_modes:
                addr_parts.append(lines[j])
                j += 1
            data["ship_to_address"] = " ".join(addr_parts).strip()
            break

    # ── 4. Date
    for line in lines:
        if date_re.match(line):
            data["date"] = line
            break

    # ── 5. Ship mode
    for line in lines:
        if line in ship_modes:
            data["ship_mode"] = line
            break

    # ── 6. Balance Due  (the $X.XX that appears BEFORE the label block)
    #    It is the money line immediately before the "Date" label line.
    for i, line in enumerate(lines):
        if line == "Date" and i > 0 and money_re.match(lines[i - 1]):
            data["balance_due"] = lines[i - 1]
            break

    # ── 7. Find the table section
    #    Locate the "Amount" header line (last of Item/Quantity/Rate/Amount)
    #    then parse items until we hit the summary dollar section.
    table_start = None
    for i, line in enumerate(lines):
        if line == "Amount" and i >= 3:
            # Verify the 3 lines before are Rate, Quantity, Item
            if lines[i-1] == "Rate" and lines[i-2] == "Quantity" and lines[i-3] == "Item":
                table_start = i + 1
                break

    if table_start is not None:
        data["line_items"] = _parse_items_block(lines, table_start, money_re, qty_re)
    else:
        data["line_items"] = [{"description": "(see document)", "sku_or_category": "",
                               "quantity": "", "unit_price": "", "amount": ""}]

    # ── 8. Summary financials
    #    After the last line item's sku_or_category line, the pattern is:
    #    $subtotal, [$discount], $shipping, $total  — all as consecutive money lines
    #    then follow: "Subtotal", ":", ["Discount (X%)"], [":"], "Shipping", ":" ...
    #
    #    Strategy: find the "Subtotal" label, look backwards for consecutive money lines.
    for i, line in enumerate(lines):
        if line == "Subtotal" and i + 1 < n and lines[i + 1] == ":":
            # Count consecutive discount labels
            has_discount = False
            disc_rate   = ""
            for k in range(i + 2, min(i + 6, n)):
                m = re.match(r'^Discount\s*\((\d+)%\)$', lines[k], re.IGNORECASE)
                if m:
                    has_discount = True
                    disc_rate = m.group(1) + "%"
                    break

            # Gather money lines immediately BEFORE "Subtotal"
            money_before = []
            j = i - 1
            while j >= 0 and money_re.match(lines[j]):
                money_before.insert(0, lines[j])
                j -= 1

            # money_before order: [subtotal, discount (if any), shipping, total]
            # But if amounts come in this order:
            # $subtotal  $discount(opt)  $shipping  $total
            if has_discount:
                # 4 values: subtotal, discount, shipping, total
                if len(money_before) >= 4:
                    data["subtotal"]      = money_before[0]
                    data["discount"]      = money_before[1]
                    data["shipping_cost"] = money_before[2]
                    data["total"]         = money_before[3]
                elif len(money_before) == 3:
                    data["subtotal"]      = money_before[0]
                    data["discount"]      = money_before[1]
                    data["total"]         = money_before[2]
                elif len(money_before) == 2:
                    data["subtotal"]      = money_before[0]
                    data["discount"]      = money_before[1]
            else:
                # 3 values: subtotal, shipping, total  OR  2: subtotal, total
                if len(money_before) >= 3:
                    data["subtotal"]      = money_before[0]
                    data["shipping_cost"] = money_before[1]
                    data["total"]         = money_before[2]
                elif len(money_before) == 2:
                    data["subtotal"]      = money_before[0]
                    data["total"]         = money_before[1]
                elif len(money_before) == 1:
                    data["subtotal"]      = money_before[0]
            break

    # ── 9. Notes
    for i, line in enumerate(lines):
        if line == "Notes" and i + 1 < n and lines[i + 1] == ":" and i + 2 < n:
            data["notes"] = lines[i + 2]
            break

    # ── 10. Terms
    for i, line in enumerate(lines):
        if line == "Terms" and i + 1 < n and lines[i + 1] == ":" and i + 2 < n:
            val = lines[i + 2]
            if not val.startswith("Order ID"):
                data["payment_terms"] = val
            break

    # ── 11. Order ID
    for line in lines:
        m = re.match(r'^Order\s+ID\s*:\s*(.+)$', line, re.IGNORECASE)
        if m:
            data["order_id"] = m.group(1).strip()
            break

    return data


def _parse_items_block(lines: list, start: int, money_re, qty_re) -> list:
    """
    Parse line items from the table section.

    Each item occupies exactly 4 consecutive lines:
        description
        qty          (integer)
        $unit_price
        $amount
    Followed by exactly 1 sku/category line.

    The block ends when we hit a label keyword (Subtotal, Shipping…).
    """
    STOP_WORDS = {"Subtotal", "Shipping", "Total", "Discount", "Notes", "Terms", "Tax"}
    items = []
    i = start
    n = len(lines)

    while i < n:
        line = lines[i]

        # Stop if we reach a summary keyword or colon-only line
        if line in STOP_WORDS or line == ":":
            break
        # Stop if this is a money line without a preceding description
        # (indicates we've entered the summary $ block)
        if money_re.match(line):
            break

        # Otherwise treat as description
        description = line
        quantity    = ""
        unit_price  = ""
        amount      = ""
        sku         = ""

        # Consume qty (next line = small integer)
        if i + 1 < n and qty_re.match(lines[i + 1]):
            quantity = lines[i + 1]
            i += 2
        else:
            i += 1
            # No numeric qty found

        # Consume unit_price (next = $X.XX)
        if i < n and money_re.match(lines[i]):
            unit_price = lines[i]
            i += 1

        # Consume amount (next = $X.XX)
        if i < n and money_re.match(lines[i]):
            amount = lines[i]
            i += 1

        # Consume sku_or_category (next non-money, non-qty, non-stop line)
        if i < n:
            peek = lines[i]
            if (peek not in STOP_WORDS and peek != ":"
                    and not money_re.match(peek) and not qty_re.match(peek)):
                sku = peek
                i += 1

        items.append({
            "description":     description,
            "sku_or_category": sku,
            "quantity":        quantity,
            "unit_price":      unit_price,
            "amount":          amount,
        })

    return items if items else [
        {"description": "(see document)", "sku_or_category": "",
         "quantity": "", "unit_price": "", "amount": ""}
    ]


# ─── OLLAMA FALLBACK (improved, tighter prompt) ────────────────────────────────
# Tight schema prompt specifically for SuperStore invoices
_OLLAMA_PROMPT = """\
You are a data extraction assistant. Extract data from the invoice text below.
Return ONLY a valid JSON object with EXACTLY these keys (use "" for missing values):

{
  "reference_number": "invoice number after # symbol",
  "order_id": "Order ID value",
  "date": "document date",
  "client_name": "Bill To name",
  "ship_to_address": "Ship To address",
  "ship_mode": "shipping method (First Class / Second Class / Standard Class / Same Day)",
  "subtotal": "Subtotal dollar amount with $ sign",
  "discount": "Discount dollar amount with $ sign, or empty",
  "shipping_cost": "Shipping dollar amount with $ sign",
  "tax": "Tax dollar amount with $ sign, or empty",
  "total": "Total dollar amount with $ sign",
  "balance_due": "Balance Due dollar amount with $ sign",
  "notes": "Notes field text",
  "payment_terms": "Terms field text",
  "line_items": [
    {
      "description": "product name only (NO category/SKU info)",
      "sku_or_category": "category/SKU line that appears below product name",
      "quantity": "numeric quantity",
      "unit_price": "Rate dollar amount with $ sign",
      "amount": "line total dollar amount with $ sign"
    }
  ]
}

RULES:
- line_items must only contain ACTUAL purchased products.
- DO NOT put Subtotal, Discount, Shipping, Tax, or Total as line items.
- The category/SKU line (e.g. "Chairs, Furniture, FUR-CH-4421") goes in sku_or_category of the SAME item, NOT as a new item.
- Return ONLY the JSON. No explanation. No markdown.

INVOICE TEXT:
"""


def process_with_ollama(file_path: Path) -> dict:
    log.info(f"Extracting text for Ollama from '{file_path.name}'…")
    text = extract_text_from_pdf(file_path)

    if not text.strip():
        raise ValueError("PDF has no extractable text (scanned image?).")

    # First try fast regex parser — only fall back to Ollama if it fails
    try:
        result = parse_invoice_regex(text)
        if result.get("reference_number") or result.get("client_name"):
            log.info("  → Regex parser succeeded, skipping Ollama.")
            return result
    except Exception:
        pass

    log.info(f"  → Sending to Ollama ({OLLAMA_MODEL})…")
    prompt = _OLLAMA_PROMPT + text[:5000]
    resp = requests.post(OLLAMA_URL, json={
        "model":   OLLAMA_MODEL,
        "prompt":  prompt,
        "stream":  False,
        "format":  "json",
        "options": {"temperature": 0.0, "top_p": 0.9},
    }, timeout=180)

    if resp.status_code != 200:
        raise ConnectionError(f"Ollama HTTP {resp.status_code}: {resp.text[:200]}")

    raw = extract_json(resp.json().get("response", "{}"))
    # Merge Ollama result with defaults to ensure all keys present
    merged = {
        "document_type": "Invoice",
        "vendor_name":   "SuperStore",
        "vendor_address": "",
        "vendor_contact": "",
        "client_address": "",
        "currency":      "USD",
        "due_date":      "",
        "tax_rate":      "",
        **raw,
    }
    return merged


# ─── GEMINI FALLBACK ───────────────────────────────────────────────────────────
_GEMINI_PROMPT = """\
Extract data from this SuperStore invoice PDF.
Return ONLY a raw JSON object with these exact keys:

{
  "reference_number": "invoice number after # symbol",
  "order_id": "Order ID value",
  "date": "document date",
  "vendor_name": "SuperStore",
  "client_name": "Bill To name",
  "ship_to_address": "Ship To address",
  "ship_mode": "shipping method",
  "currency": "USD",
  "subtotal": "Subtotal with $ sign",
  "discount": "Discount with $ sign or empty",
  "shipping_cost": "Shipping with $ sign",
  "tax": "Tax with $ sign or empty",
  "tax_rate": "Tax % if shown or empty",
  "total": "Total with $ sign",
  "balance_due": "Balance Due with $ sign",
  "payment_terms": "Terms field",
  "notes": "Notes field",
  "line_items": [
    {
      "description": "product name only",
      "sku_or_category": "category/SKU line below product name",
      "quantity": "qty",
      "unit_price": "Rate with $ sign",
      "amount": "line total with $ sign"
    }
  ]
}

STRICT RULES:
- DO NOT add Subtotal/Discount/Shipping/Tax/Total as line_items.
- sku_or_category for each item is the sub-line like "Chairs, Furniture, FUR-CH-4421".
- Use "" for any missing field, never null.
- Return ONLY the JSON."""


def process_with_gemini(file_path: Path) -> dict:
    try:
        import google.generativeai as genai
    except ImportError:
        raise ImportError("google-generativeai not installed. Run: pip install google-generativeai")

    genai.configure(api_key=GEMINI_API_KEY)
    log.info(f"Uploading '{file_path.name}' to Gemini…")
    uploaded_file = None
    try:
        uploaded_file = genai.upload_file(path=str(file_path), mime_type="application/pdf")
        model = genai.GenerativeModel(
            'gemini-2.5-flash',
            generation_config={"response_mime_type": "application/json"}
        )
        max_retries = 4
        for attempt in range(max_retries):
            try:
                time.sleep(4)
                response = model.generate_content([uploaded_file, _GEMINI_PROMPT])
                return extract_json(response.text)
            except Exception as e:
                err_str = str(e)
                is_rate = "429" in err_str or "quota" in err_str.lower() or "exhausted" in err_str.lower()
                if is_rate and attempt < max_retries - 1:
                    wait = 65 * (attempt + 1)
                    log.warning(f"Rate limit — waiting {wait}s (retry {attempt+1}/{max_retries})…")
                    time.sleep(wait)
                    continue
                raise
    finally:
        if uploaded_file is not None:
            try:
                genai.delete_file(uploaded_file.name)
            except Exception:
                pass


# ─── JSON EXTRACTION HELPER ───────────────────────────────────────────────────
def extract_json(text: str) -> dict:
    text = re.sub(r'^```(?:json)?\s*', '', text.strip(), flags=re.IGNORECASE)
    text = re.sub(r'\s*```$', '', text.strip())
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    raise ValueError(f"Could not parse JSON from model output:\n{text[:300]}")


# ─── FILE STABLE WAIT ─────────────────────────────────────────────────────────
def wait_for_file_stable(path: Path, timeout: int = 30) -> bool:
    deadline = time.time() + timeout
    last_size = -1
    while time.time() < deadline:
        try:
            size = path.stat().st_size
        except FileNotFoundError:
            return False
        if size == last_size and size > 0:
            return True
        last_size = size
        time.sleep(0.5)
    return False


# ─── DATABASE COLUMNS (clean, minimal, no duplicates) ─────────────────────────
HEADERS = [
    'Timestamp', 'File Name',
    'Invoice #', 'Order ID', 'Date',
    'Vendor', 'Client', 'Ship To Address', 'Ship Mode',
    'Currency', 'Subtotal', 'Discount', 'Shipping', 'Tax', 'Tax Rate', 'Total', 'Balance Due',
    'Payment Terms', 'Notes',
    'Item Description', 'SKU / Category', 'Qty', 'Unit Price', 'Line Amount',
]

COL_WIDTHS = [
    18, 30,
    12, 28, 14,
    20, 24, 35, 16,
    10, 12, 12, 12, 10, 10, 12, 12,
    16, 30,
    40, 28, 6, 12, 14,
]

_csv_lock  = threading.Lock()
_xlsx_lock = threading.Lock()


def append_to_database(file_name: str, data: dict):
    timestamp  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line_items = data.get("line_items") or [{}]

    # Consolidate line items into pipe-separated strings
    desc_list  = []
    sku_list   = []
    qty_list   = []
    price_list = []
    amt_list   = []

    for item in line_items:
        desc = str(item.get("description", "")).strip()
        if desc and desc != "(see document)":
            desc_list.append(desc)
            sku_list.append(str(item.get("sku_or_category", "")).strip())
            qty_list.append(str(item.get("quantity", "")).strip())
            price_list.append(str(item.get("unit_price", "")).strip())
            amt_list.append(str(item.get("amount", "")).strip())

    row = [
        timestamp,
        file_name,
        data.get("reference_number", ""),
        data.get("order_id", ""),
        data.get("date", ""),
        data.get("vendor_name", "SuperStore"),
        data.get("client_name", ""),
        data.get("ship_to_address", ""),
        data.get("ship_mode", ""),
        data.get("currency", "USD"),
        data.get("subtotal", ""),
        data.get("discount", ""),
        data.get("shipping_cost", ""),
        data.get("tax", ""),
        data.get("tax_rate", ""),
        data.get("total", ""),
        data.get("balance_due", ""),
        data.get("payment_terms", ""),
        data.get("notes", ""),
        " | ".join(desc_list),
        " | ".join(sku_list),
        " | ".join(qty_list),
        " | ".join(price_list),
        " | ".join(amt_list),
    ]

    with _csv_lock:
        with open(DATABASE_FILE, 'a', newline='', encoding='utf-8') as f:
            csv.writer(f).writerow(row)


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────
def update_excel_file():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not DATABASE_FILE.exists() or DATABASE_FILE.stat().st_size == 0:
            return

        with _csv_lock:
            with open(DATABASE_FILE, 'r', encoding='utf-8') as f:
                rows = list(csv.reader(f))

        if len(rows) < 2:
            return

        with _xlsx_lock:
            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Data"

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            thin        = Side(border_style="thin", color="D0D0D0")
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = cell_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"

            alt_fill  = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")
            data_font = Font(name="Calibri", size=9)
            money_cols = {col_idx for col_idx, h in enumerate(HEADERS, 1)
                          if h in ('Subtotal', 'Discount', 'Shipping', 'Tax',
                                   'Total', 'Balance Due', 'Unit Price', 'Line Amount')}

            for row_idx, row_data in enumerate(rows[1:], start=2):
                for col_idx in range(1, len(HEADERS) + 1):
                    value = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font   = data_font
                    cell.border = cell_border
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill
                    if col_idx in money_cols:
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.auto_filter.ref = ws.dimensions
            wb.save(EXCEL_FILE)
            log.info(f"Excel updated → {EXCEL_FILE.name}  ({len(rows)-1} rows)")

    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
    except Exception as exc:
        log.error(f"Excel update failed: {exc}")


# ─── SETUP ────────────────────────────────────────────────────────────────────
def setup_environment():
    load_config()
    for d in [INBOX_DIR, PROCESSED_DIR, FAILED_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    if not DATABASE_FILE.exists():
        with open(DATABASE_FILE, 'w', newline='', encoding='utf-8') as f:
            csv.writer(f).writerow(HEADERS)

    if ENGINE == 'gemini' and is_configured():
        try:
            import google.generativeai as genai
            genai.configure(api_key=GEMINI_API_KEY)
            log.info("Gemini API configured.")
        except ImportError:
            log.warning("google-generativeai not installed. pip install google-generativeai")


# ─── DUPLICATE DETECTION ──────────────────────────────────────────────────────
def is_duplicate(file_name: str) -> bool:
    """
    Return True if *file_name* already has a row in Extracted_Database.csv
    OR already exists in the Processed/ folder.
    Both checks are needed because the CSV may be cleared without moving
    files back, or vice-versa.
    """
    # 1. Check Processed/ directory
    if (PROCESSED_DIR / file_name).exists():
        return True
    # 2. Check CSV (column index 1 = 'File Name')
    if not DATABASE_FILE.exists():
        return False
    try:
        with _csv_lock:
            with open(DATABASE_FILE, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader, None)          # skip header
                for row in reader:
                    if len(row) > 1 and row[1].strip() == file_name.strip():
                        return True
    except Exception:
        pass
    return False


# ─── MAIN FILE HANDLER ────────────────────────────────────────────────────────
def handle_new_file(file_path: Path):
    file_path = Path(file_path)
    if file_path.suffix.lower() != '.pdf' or not file_path.exists():
        return

    # ── Duplicate guard ────────────────────────────────────────────────────────
    if is_duplicate(file_path.name):
        log.warning(f"DUPLICATE — '{file_path.name}' already processed. Skipping & removing from Inbox.")
        audit.log_event("FILE_DUPLICATE_SKIPPED", file=file_path.name,
                        reason="Already in Processed/ or CSV database")
        try:
            file_path.unlink()          # remove from Inbox so it doesn't accumulate
        except Exception:
            pass
        return

    log.info(f"─── Processing: {file_path.name} ───")
    audit.log_event("FILE_PROCESSING_START", file=file_path.name, engine=ENGINE)

    if not wait_for_file_stable(file_path):
        reason = "Timeout waiting for file to become stable"
        log.error(f"Timeout waiting for file: {file_path.name}")
        audit.extraction_fail(file=file_path.name, error=reason)
        _move_to(file_path, FAILED_DIR)
        notify_human_review(file_name=file_path.name, reason=reason)
        return

    global currently_processing
    with processing_lock:
        currently_processing = file_path.name

    try:
        if not is_configured():
            reason = "Bot not configured — engine/api_key missing in config.json"
            log.error(f"Not configured — skipping '{file_path.name}'. Set engine/api_key in config.json.")
            audit.log_event("FILE_SKIPPED", file=file_path.name, reason=reason)
            with processing_lock:
                currently_processing = None
            return

        # ── Choose extraction method
        if ENGINE == 'gemini':
            data = process_with_gemini(file_path)
        elif ENGINE == 'ollama':
            data = process_with_ollama(file_path)
        else:
            # Default: pure regex parser (fastest, most accurate for SuperStore PDFs)
            text = extract_text_from_pdf(file_path)
            if not text.strip():
                raise ValueError("No text extracted — is this a scanned image PDF?")
            data = parse_invoice_regex(text)

        append_to_database(file_path.name, data)
        update_excel_file()
        _move_to(file_path, PROCESSED_DIR)

        n_items = len([i for i in (data.get("line_items") or [])
                       if i.get("description") and i["description"] != "(see document)"])
        log.info(f"✔ Done: '{file_path.name}' | Invoice #{data.get('reference_number','')} "
                 f"| {n_items} line item(s) | Total: {data.get('total','')}\n")
        audit.extraction_ok(
            file=file_path.name,
            engine=ENGINE,
            invoice=data.get("reference_number", ""),
            total=data.get("total", ""),
            items=n_items,
        )

    except Exception as exc:
        reason = str(exc)
        log.error(f"✘ Failed '{file_path.name}': {exc}\n")
        audit.extraction_fail(file=file_path.name, error=reason)
        _move_to(file_path, FAILED_DIR)
        notify_human_review(file_name=file_path.name, reason=reason)

    finally:
        with processing_lock:
            currently_processing = None
        if ENGINE == 'gemini' and is_configured():
            time.sleep(INTER_FILE_DELAY)


def _move_to(src: Path, dest_dir: Path):
    dest = dest_dir / src.name
    if dest.exists():
        dest = dest_dir / f"{src.stem}_{int(time.time())}{src.suffix}"
    try:
        shutil.move(str(src), str(dest))
    except Exception as exc:
        log.error(f"Could not move {src.name} → {dest_dir.name}: {exc}")


# ─── WORKER THREAD ────────────────────────────────────────────────────────────
def _queue_worker():
    while True:
        file_path = _file_queue.get()
        if file_path is None:
            break
        handle_new_file(file_path)
        _file_queue.task_done()


def enqueue_file(file_path):
    p = Path(file_path)
    audit.file_queued(file=p.name)
    _file_queue.put(p)


# ─── WATCHDOG ─────────────────────────────────────────────────────────────────
class PDFHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory and event.src_path.lower().endswith('.pdf'):
            log.info(f"Watchdog: new file → {Path(event.src_path).name}")
            enqueue_file(event.src_path)


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────
def start_bot():
    global _worker_thread
    setup_environment()

    log.info("═══════════════════════════════════════════")
    log.info("  Invoice RPA Bot  —  Ready")
    log.info(f"  Engine  : {ENGINE.upper()}")
    log.info(f"  Inbox   : {INBOX_DIR}")
    log.info(f"  Output  : {EXCEL_FILE.name}  +  {DATABASE_FILE.name}")
    log.info("  Drop PDFs into the Inbox folder to start…")
    log.info("═══════════════════════════════════════════")

    _worker_thread = threading.Thread(target=_queue_worker, name="rpa-worker", daemon=True)
    _worker_thread.start()

    existing = sorted(INBOX_DIR.glob("*.pdf"))
    if existing:
        log.info(f"Found {len(existing)} PDF(s) already in Inbox — queuing…")
    for f in existing:
        enqueue_file(f)

    handler  = PDFHandler()
    observer = Observer()
    observer.schedule(handler, str(INBOX_DIR), recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        _file_queue.put(None)
        log.info("RPA Bot stopped.")
    observer.join()


if __name__ == "__main__":
    start_bot()